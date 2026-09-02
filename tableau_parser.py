"""
Tableau Figure Extractor
========================
Reads a .twbx file, queries the embedded .hyper extract, and presents
figures from each worksheet interactively. You choose which figures to
include in the output.

Outputs:
  - figures.xlsx  — formatted Excel workbook
  - figures.html  — clean HTML dashboard

Usage:
    python tableau_figure_extractor.py --file /path/to/workbook.twbx

Requirements:
    pip install tableauhyperapi openpyxl
"""

import sys, os, re, json, zipfile, argparse, tempfile
from pathlib import Path
from collections import defaultdict
import xml.etree.ElementTree as ET

try:
    from tableauhyperapi import HyperProcess, Telemetry, Connection, TableName
    HAS_HYPER = True
except ImportError:
    HAS_HYPER = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ─────────────────────────────────────────────
#  NAMESPACE STRIPPING
# ─────────────────────────────────────────────

def _strip_namespaces(content: bytes) -> bytes:
    content = re.sub(rb'\s+xmlns(?::\w+)?="[^"]*"', b'', content)
    content = re.sub(rb"\s+xmlns(?::\w+)?='[^']*'", b'', content)
    content = re.sub(rb'<(/?)[\w][\w.-]*:([\w][\w.-]*)', rb'<\1\2', content)
    content = re.sub(rb' [\w][\w.-]*:([\w][\w.-]*)=', rb' \1=', content)
    return content


# ─────────────────────────────────────────────
#  EXTRACT .twbx
# ─────────────────────────────────────────────

def extract_twbx(filepath: Path, dest: Path):
    """
    Unzip a .twbx into dest. Returns (twb_path, hyper_path).
    hyper_path may be None if no extract is found.
    """
    with zipfile.ZipFile(filepath, "r") as zf:
        zf.extractall(dest)
        names = zf.namelist()

    twb_files   = [dest / n for n in names if n.endswith(".twb")]
    hyper_files = [dest / n for n in names if n.endswith(".hyper")]

    if not twb_files:
        raise ValueError("No .twb found inside the .twbx archive.")

    twb_path   = twb_files[0]
    hyper_path = hyper_files[0] if hyper_files else None
    return twb_path, hyper_path


# ─────────────────────────────────────────────
#  PARSE WORKBOOK XML
# ─────────────────────────────────────────────

def clean_field_name(raw: str) -> str:
    """
    Turn Tableau internal field references into readable labels.
    e.g. '[Sum:Sales:qk]' -> 'Sum of Sales'
         '[Sales]'        -> 'Sales'
         'SUM([Revenue])' -> 'Sum of Revenue'
    """
    s = raw.strip()

    # Pattern: [Agg:FieldName:suffix] e.g. [Sum:Sales:qk]
    m = re.match(r'\[(\w+):([^\]:]+)(?::[^\]]+)?\]', s)
    if m:
        agg, field = m.group(1), m.group(2)
        agg_map = {
            "sum": "Sum of", "avg": "Avg of", "count": "Count of",
            "countd": "Count Distinct of", "min": "Min of", "max": "Max of",
            "median": "Median of", "attr": "",
        }
        prefix = agg_map.get(agg.lower(), agg.title()+" of")
        field = field.replace("_", " ").replace("-", " ").title()
        return f"{prefix} {field}".strip() if prefix else field

    # Pattern: SUM([Field]) or AVG([Field])
    m = re.match(r'(\w+)\(\[([^\]]+)\]\)', s)
    if m:
        agg, field = m.group(1), m.group(2)
        agg_map = {
            "SUM": "Sum of", "AVG": "Avg of", "COUNT": "Count of",
            "COUNTD": "Count Distinct of", "MIN": "Min of", "MAX": "Max of",
            "MEDIAN": "Median of", "ATTR": "",
        }
        prefix = agg_map.get(agg.upper(), agg.title()+" of")
        field = field.replace("_", " ").replace("-", " ").title()
        return f"{prefix} {field}".strip() if prefix else field

    # Plain [FieldName]
    m = re.match(r'\[([^\]]+)\]', s)
    if m:
        return m.group(1).replace("_", " ").replace("-", " ").title()

    return s.replace("_", " ").replace("-", " ").title()


def parse_worksheet_fields(twb_path: Path):
    """
    Parse the .twb XML and return a dict:
      { worksheet_name: [{"raw": ..., "label": ..., "role": "measure|dimension"}] }
    """
    content = _strip_namespaces(twb_path.read_bytes())
    root = ET.fromstring(content)

    # Build a lookup of column captions from the datasource
    caption_map = {}
    for col in root.findall(".//column"):
        name    = col.get("name", "")
        caption = col.get("caption", "")
        if name and caption:
            caption_map[name] = caption

    ws_fields = {}
    for ws in root.findall("worksheets/worksheet"):
        ws_name = ws.get("name", "Unknown")
        fields  = []
        seen    = set()

        # Walk all encodings (rows, columns, marks, filters) to collect fields
        for enc in ws.findall(".//encoding"):
            field = enc.get("field", "")
            if not field or field in seen:
                continue
            seen.add(field)
            # Skip internal Tableau fields
            if field.startswith("[Tableau") or field.startswith("[Number of Records]"):
                continue
            label = caption_map.get(field, clean_field_name(field))
            role  = "measure" if any(
                agg in field.lower() for agg in ["sum:", "avg:", "count:", "min:", "max:"]
            ) else "dimension"
            fields.append({"raw": field, "label": label, "role": role})

        # Also check rows/cols shelves for measure fields
        for shelf in ["rows", "cols"]:
            el = ws.find(f"table/view/{shelf}")
            if el is not None and el.text:
                for part in re.findall(r'\[[^\]]+(?::[^\]]+)*\]', el.text):
                    if part in seen:
                        continue
                    seen.add(part)
                    if part.startswith("[Tableau") or "Number of Records" in part:
                        continue
                    label = caption_map.get(part, clean_field_name(part))
                    role  = "measure" if any(
                        agg in part.lower() for agg in ["sum:", "avg:", "count:", "min:", "max:", "sum(", "avg("]
                    ) else "dimension"
                    fields.append({"raw": part, "label": label, "role": role})

        ws_fields[ws_name] = fields

    return ws_fields


# ─────────────────────────────────────────────
#  QUERY .hyper FILE
# ─────────────────────────────────────────────

def get_hyper_tables(hyper_path: Path):
    """Return list of (schema, table_name) from the hyper file."""
    tables = []
    with HyperProcess(telemetry=Telemetry.DO_NOT_SEND_USAGE_DATA_TO_TABLEAU) as hyper:
        with Connection(hyper.endpoint, str(hyper_path)) as conn:
            catalog = conn.catalog
            for schema in catalog.get_schema_names():
                for tbl in catalog.get_table_names(schema):
                    tables.append((str(schema), str(tbl.name)))
    return tables


def get_hyper_schema(hyper_path: Path) -> list:
    """
    Open the .hyper file and return a list of table dicts:
      [{ "table_obj": ..., "schema": str, "table": str, "columns": [{"name":str,"type":str}] }]
    """
    tables = []
    with HyperProcess(telemetry=Telemetry.DO_NOT_SEND_USAGE_DATA_TO_TABLEAU) as hyper:
        with Connection(hyper.endpoint, str(hyper_path)) as conn:
            catalog = conn.catalog
            for schema in catalog.get_schema_names():
                for tbl in catalog.get_table_names(schema):
                    td   = catalog.get_table_definition(tbl)
                    cols = []
                    for c in td.columns:
                        type_str = str(c.type).lower()
                        cols.append({"name": c.name.unescaped, "type": type_str})
                    tables.append({
                        "table_obj": tbl,
                        "schema":    str(schema),
                        "table":     str(tbl.name),
                        "columns":   cols,
                    })
    return tables


def is_numeric_type(type_str: str) -> bool:
    """Return True if this Hyper column type is numeric."""
    return any(t in type_str for t in [
        "int", "float", "double", "numeric", "decimal",
        "real", "big", "small", "tiny", "money", "currency",
    ])


def query_all_figures(hyper_path: Path, ws_names: list) -> dict:
    """
    Read every column from the .hyper extract and return the first
    non-null value from each one — exactly as Tableau already computed it.
    No aggregation applied.

    Returns:
      { worksheet_name: [{"label": str, "value": any}] }
    All worksheets share the same list; the user assigns figures to
    worksheets interactively.
    """
    schema = get_hyper_schema(hyper_path)
    if not schema:
        return {}

    primary  = schema[0]
    tbl_obj  = primary["table_obj"]
    all_cols = primary["columns"]

    figures = []

    with HyperProcess(telemetry=Telemetry.DO_NOT_SEND_USAGE_DATA_TO_TABLEAU) as hyper:
        with Connection(hyper.endpoint, str(hyper_path)) as conn:
            for col in all_cols:
                col_name  = col["name"]
                safe_name = f'"{col_name}"'
                display   = col_name.replace("_", " ").replace("-", " ").title()
                try:
                    # Grab the first non-null value — already the computed figure
                    sql    = f'SELECT {safe_name} FROM {tbl_obj} WHERE {safe_name} IS NOT NULL LIMIT 1'
                    result = conn.execute_scalar_query(sql)
                    if result is not None:
                        figures.append({"label": display, "value": result})
                except Exception:
                    continue

    if not figures:
        return {}

    return {ws: figures for ws in ws_names}



# ─────────────────────────────────────────────
#  FORMAT VALUES
# ─────────────────────────────────────────────

def fmt(value):
    """Format a value for display."""
    if value is None:
        return "—"
    if isinstance(value, float):
        if value == int(value):
            return f"{int(value):,}"
        return f"{value:,.2f}"
    if isinstance(value, int):
        return f"{value:,}"
    return str(value)


# ─────────────────────────────────────────────
#  INTERACTIVE SELECTION
# ─────────────────────────────────────────────

def interactive_select(figures_by_ws: dict) -> list:
    """
    New flow — figures come from the .hyper schema directly, not the XML.
    All worksheets share the same available figure list.

    Step 1: Show all available figures and let the user pick which ones to include.
    Step 2: For each picked figure, ask which worksheet to label it under
            (or let them type a custom label if they prefer).

    Returns a list of { "worksheet": str, "label": str, "value": ..., "value_fmt": str }
    """
    if not figures_by_ws:
        print("\nNo figures available.")
        return []

    # All worksheets share the same figure list — grab it from the first entry
    all_figs = next(iter(figures_by_ws.values()))
    ws_names = list(figures_by_ws.keys())

    print("\n" + "═"*60)
    print("  STEP 1 — AVAILABLE FIGURES")
    print("  These figures come directly from the extract.")
    print("  Enter the numbers you want to include (e.g. 1,3,5)")
    print("  or press ENTER to include all.")
    print("═"*60)
    print()
    for i, fig in enumerate(all_figs, start=1):
        print(f"  [{i:>3}] {fig['label']:<40} {fmt(fig['value']):>15}")
    print()

    while True:
        raw = input("  Select figures (ENTER = all): ").strip()
        if not raw:
            chosen_figs = list(range(len(all_figs)))
            break
        try:
            choices = [int(x.strip()) - 1 for x in raw.split(",")]
            if all(0 <= c < len(all_figs) for c in choices):
                chosen_figs = choices
                break
            else:
                print(f"  Please enter numbers between 1 and {len(all_figs)}.")
        except ValueError:
            print("  Invalid input — enter numbers separated by commas.")

    if not chosen_figs:
        return []

    print(f"\n  {len(chosen_figs)} figure(s) selected.")

    # Step 2 — assign each figure to a worksheet
    print("\n" + "═"*60)
    print("  STEP 2 — ASSIGN TO WORKSHEETS")
    print("  For each figure, choose which worksheet to label it under.")
    print("  Enter a number from the list, or type a custom name.")
    print("═"*60)

    print("\n  Worksheets:")
    for i, ws in enumerate(ws_names, start=1):
        print(f"  [{i:>3}] {ws}")
    print()

    selected = []
    for idx in chosen_figs:
        fig = all_figs[idx]
        print(f"  Figure: {fig['label']}  =  {fmt(fig['value'])}")
        while True:
            raw = input("  Worksheet (number or custom name, ENTER to skip): ").strip()
            if not raw:
                break
            # Check if it's a number
            try:
                ws_idx = int(raw) - 1
                if 0 <= ws_idx < len(ws_names):
                    ws_label = ws_names[ws_idx]
                else:
                    print(f"  Please enter a number between 1 and {len(ws_names)}.")
                    continue
            except ValueError:
                # Treat as a custom worksheet name
                ws_label = raw

            selected.append({
                "worksheet": ws_label,
                "label":     fig["label"],
                "value":     fig["value"],
                "value_fmt": fmt(fig["value"]),
            })
            break
        print()

    return selected


# ─────────────────────────────────────────────
#  EXCEL OUTPUT
# ─────────────────────────────────────────────

_HEADER_BG = "1A3D2B"
_HEADER_FG = "6EE7B7"
_ROW_ALT   = "F0FDF4"
_WHITE     = "FFFFFF"
_DARK      = "1F2937"
_GREEN     = "34D399"
_BLUE      = "4F8EF7"

def _thin():
    s = Side(style="thin", color="D1D5DB")
    return Border(left=s, right=s, top=s, bottom=s)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def write_figures_xlsx(selected: list, workbook_name: str, path: Path):
    if not HAS_OPENPYXL:
        print("  ⚠  openpyxl not installed. Run: pip install openpyxl")
        return

    wb  = Workbook()
    ws1 = wb.active
    ws1.title = "Selected Figures"

    headers = ["Workbook", "Worksheet", "Figure", "Value"]
    ws1.append(headers)
    for col in range(1, 5):
        cell = ws1.cell(row=1, column=col)
        cell.font      = Font(name="Arial", bold=True, color=_HEADER_FG, size=10)
        cell.fill      = _fill(_HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = _thin()
    ws1.freeze_panes = "A2"

    for i, row in enumerate(selected, start=2):
        row_data = [row.get("workbook", workbook_name), row["worksheet"], row["label"], row["value_fmt"]]
        ws1.append(row_data)
        rf = _fill(_ROW_ALT) if i % 2 == 0 else _fill(_WHITE)
        for col, val in enumerate(row_data, start=1):
            cell = ws1.cell(row=i, column=col)
            cell.border    = _thin()
            cell.alignment = Alignment(
                horizontal="right" if col == 4 else "left",
                vertical="center"
            )
            cell.fill = rf
            cell.font = Font(
                name="Arial", bold=(col <= 2), color=_DARK, size=9
            )

    # Auto-fit columns
    for col_cells in ws1.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws1.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 60)

    # Summary sheet — group by worksheet
    ws2 = wb.create_sheet("By Worksheet")
    ws2.append(["Worksheet", "Figure", "Value"])
    for col in range(1, 4):
        cell = ws2.cell(row=1, column=col)
        cell.font      = Font(name="Arial", bold=True, color=_HEADER_FG, size=10)
        cell.fill      = _fill(_HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = _thin()
    ws2.freeze_panes = "A2"

    for i, row in enumerate(selected, start=2):
        row_data = [row["worksheet"], row["label"], row["value_fmt"]]
        ws2.append(row_data)
        rf = _fill(_ROW_ALT) if i % 2 == 0 else _fill(_WHITE)
        for col, val in enumerate(row_data, start=1):
            cell = ws2.cell(row=i, column=col)
            cell.border    = _thin()
            cell.alignment = Alignment(
                horizontal="right" if col == 3 else "left",
                vertical="center"
            )
            cell.fill = rf
            cell.font = Font(name="Arial", bold=(col == 1), color=_DARK, size=9)

    for col_cells in ws2.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws2.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 60)

    wb.save(path)
    print(f"  ✓ XLSX → {path}")


# ─────────────────────────────────────────────
#  HTML OUTPUT
# ─────────────────────────────────────────────

def write_figures_html(selected: list, workbook_name: str, path: Path):
    # Group by workbook then worksheet
    by_wb = defaultdict(lambda: defaultdict(list))
    for row in selected:
        wb_key = row.get("workbook", workbook_name)
        by_wb[wb_key][row["worksheet"]].append(row)

    cards_html = ""
    for wb_key, sheets in by_wb.items():
        # Workbook section header (only shown if more than one workbook)
        wb_count = len(by_wb)
        if wb_count > 1:
            cards_html += f"""
      <div class="wb-section-header">📁 {wb_key}</div>"""
        for ws_name, figs in sheets.items():
            fig_html = "".join(f"""
        <div class="fig-row">
          <span class="fig-label">{fig['label']}</span>
          <span class="fig-value">{fig['value_fmt']}</span>
        </div>""" for fig in figs)

            cards_html += f"""
      <div class="card">
        <div class="card-header">
          <span class="ws-icon">📄</span>
          <span class="ws-name">{ws_name}</span>
          <span class="fig-count">{len(figs)} figure{'s' if len(figs) != 1 else ''}</span>
        </div>
        <div class="card-body">{fig_html}
        </div>
      </div>"""

    total = len(selected)
    ws_count = sum(len(s) for s in by_wb.values())
    data_json = json.dumps(selected, ensure_ascii=False, default=str)

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Figures — {workbook_name}</title>
<style>
:root{{
  --bg:#0f1117;--surface:#1a1d27;--surface2:#22263a;--border:#2e3348;
  --green:#34d399;--accent:#4f8ef7;--text:#e2e8f0;--muted:#8892a4;
  --font:'Segoe UI',system-ui,sans-serif;--mono:'Cascadia Code','Fira Code',monospace;
}}
*{{box-sizing:border-box;margin:0;padding:0}}
body{{background:var(--bg);color:var(--text);font-family:var(--font);font-size:14px;min-height:100vh}}
header{{background:linear-gradient(135deg,#111a14,#0f1117);border-bottom:1px solid var(--border);
  padding:20px 32px;display:flex;align-items:center;gap:16px}}
header h1{{font-size:20px;font-weight:700}}
header h1 span{{color:var(--green)}}
.pill{{background:var(--surface2);border:1px solid var(--border);border-radius:99px;
  padding:3px 12px;font-size:12px;color:var(--muted)}}
.toolbar{{display:flex;gap:10px;padding:16px 32px;align-items:center;border-bottom:1px solid var(--border);
  background:var(--surface);flex-wrap:wrap}}
.toolbar input{{flex:1;min-width:200px;background:var(--bg);border:1px solid var(--border);
  border-radius:7px;padding:8px 14px;color:var(--text);font-size:13px;outline:none;transition:border-color .15s}}
.toolbar input:focus{{border-color:var(--green)}}
.toolbar input::placeholder{{color:var(--muted)}}
.btn{{background:var(--green);color:#0a1f12;border:none;border-radius:7px;
  padding:8px 16px;font-size:13px;font-weight:700;cursor:pointer;transition:opacity .15s}}
.btn:hover{{opacity:.85}}
.btn.sec{{background:var(--surface2);border:1px solid var(--border);color:var(--text);font-weight:600}}
main{{padding:24px 32px;max-width:1200px}}
.stats{{display:flex;gap:12px;margin-bottom:24px;flex-wrap:wrap}}
.stat{{background:var(--surface);border:1px solid var(--border);border-radius:8px;
  padding:14px 20px;flex:1;min-width:120px}}
.stat .num{{font-size:26px;font-weight:700;color:var(--green)}}
.stat .lbl{{font-size:12px;color:var(--muted);margin-top:2px}}
.grid{{display:grid;grid-template-columns:repeat(auto-fill,minmax(360px,1fr));gap:16px}}
.card{{background:var(--surface);border:1px solid var(--border);border-radius:10px;overflow:hidden;transition:box-shadow .2s}}
.card:hover{{box-shadow:0 0 0 1px var(--green)}}
.card-header{{display:flex;align-items:center;gap:10px;padding:14px 18px;
  border-bottom:1px solid var(--border);background:var(--surface2)}}
.ws-icon{{font-size:16px}}
.ws-name{{font-weight:600;font-size:14px;flex:1;color:var(--green)}}
.fig-count{{font-size:11px;color:var(--muted);background:var(--bg);border:1px solid var(--border);
  border-radius:99px;padding:2px 8px}}
.card-body{{padding:8px 0}}
.fig-row{{display:flex;align-items:center;justify-content:space-between;
  padding:10px 18px;border-bottom:1px solid var(--border);transition:background .15s}}
.fig-row:last-child{{border-bottom:none}}
.fig-row:hover{{background:var(--surface2)}}
.fig-label{{font-size:13px;color:var(--muted);flex:1}}
.fig-value{{font-size:15px;font-weight:700;color:var(--text);font-family:var(--mono);margin-left:16px}}
.empty{{text-align:center;padding:60px;color:var(--muted)}}
.wb-section-header{{grid-column:1/-1;font-size:15px;font-weight:700;color:var(--green);
  padding:8px 4px 4px;border-bottom:1px solid var(--border);margin-bottom:4px}}
.hidden{{display:none}}
</style>
</head>
<body>
<header>
  <span style="font-size:24px">📊</span>
  <h1>Tableau <span>Figures</span></h1>
  <span class="pill">{workbook_name}</span>
</header>
<div class="toolbar">
  <input id="search" placeholder="Search worksheet or figure name…" oninput="filterCards()">
  <button class="btn sec" onclick="expandAll()">Expand All</button>
  <button class="btn" onclick="exportCSV()">⬇ Export CSV</button>
</div>
<main>
  <div class="stats">
    <div class="stat"><div class="num">{ws_count}</div><div class="lbl">Worksheets</div></div>
    <div class="stat"><div class="num">{total}</div><div class="lbl">Figures Selected</div></div>
  </div>
  <div class="grid" id="grid">
    {cards_html}
  </div>
  <div class="empty hidden" id="empty-state">No matching figures found.</div>
</main>
<script>
const DATA = {data_json};

function filterCards() {{
  const q = document.getElementById('search').value.toLowerCase().trim();
  let any = false;
  document.querySelectorAll('.card').forEach(card => {{
    const ws   = card.querySelector('.ws-name').textContent.toLowerCase();
    const rows = card.querySelectorAll('.fig-row');
    let cardMatch = ws.includes(q);
    rows.forEach(row => {{
      const lbl = row.querySelector('.fig-label').textContent.toLowerCase();
      const val = row.querySelector('.fig-value').textContent.toLowerCase();
      const match = !q || lbl.includes(q) || val.includes(q) || ws.includes(q);
      row.classList.toggle('hidden', !match && !!q);
      if (match) cardMatch = true;
    }});
    card.classList.toggle('hidden', !cardMatch && !!q);
    if (cardMatch) any = true;
  }});
  document.getElementById('empty-state').classList.toggle('hidden', any || !q);
}}

function expandAll() {{
  document.querySelectorAll('.fig-row').forEach(r => r.classList.remove('hidden'));
  document.querySelectorAll('.card').forEach(c => c.classList.remove('hidden'));
  document.getElementById('search').value = '';
  document.getElementById('empty-state').classList.add('hidden');
}}

function exportCSV() {{
  const rows = [['Worksheet','Figure','Value']];
  DATA.forEach(d => rows.push([d.worksheet, d.label, d.value_fmt]));
  const csv = rows.map(r => r.map(v => '"'+String(v).replace(/"/g,'""')+'"').join(',')).join('\\n');
  const blob = new Blob([csv], {{type:'text/csv'}});
  const url  = URL.createObjectURL(blob);
  const a    = document.createElement('a');
  a.href=url; a.download='{workbook_name}_figures.csv'; a.click();
  URL.revokeObjectURL(url);
}}
</script>
</body>
</html>"""

    with open(path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"  ✓ HTML → {path}")


# ─────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────

def process_workbook(filepath: Path) -> list:
    """
    Extract, parse, query and interactively select figures from a single .twbx.
    Returns selected figure dicts, each tagged with a 'workbook' key.
    """
    print(f"\n{'='*60}")
    print(f"  Workbook: {filepath.name}")
    print(f"{'='*60}")

    with tempfile.TemporaryDirectory() as tmpdir:
        tmp = Path(tmpdir)

        print("  Extracting archive...")
        try:
            twb_path, hyper_path = extract_twbx(filepath, tmp)
        except Exception as e:
            print(f"  Error extracting: {e} — skipping.")
            return []

        if not hyper_path:
            print("  No .hyper extract found — skipping.")
            print("  (This script requires workbooks with a data extract.)")
            return []

        print(f"  Found extract: {hyper_path.name}")

        print("  Parsing worksheet fields...")
        try:
            ws_fields = parse_worksheet_fields(twb_path)
            print(f"  Found {len(ws_fields)} worksheets")
        except Exception as e:
            print(f"  Error parsing workbook XML: {e} — skipping.")
            return []

        print("  Querying extract...")
        try:
            ws_names      = list(ws_fields.keys())
            figures_by_ws = query_all_figures(hyper_path, ws_names)
            total_figs    = len(next(iter(figures_by_ws.values()), []))
            print(f"  Found {total_figs} available figures to choose from")
        except Exception as e:
            print(f"  Error querying extract: {e} — skipping.")
            return []

        if not figures_by_ws:
            print("  No numeric columns found in extract — skipping.")
            return []

        selected = interactive_select(figures_by_ws)

        # Tag every selected figure with its workbook name
        wb_name = filepath.stem
        for row in selected:
            row["workbook"] = wb_name

        return selected


def main():
    parser = argparse.ArgumentParser(description="Tableau figure extractor.")
    group  = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--file",   help="Single .twbx file to process")
    group.add_argument("--folder", help="Folder containing .twbx files to process")
    parser.add_argument("--output", default="figures",
                        help="Output filename prefix (default: figures)")
    args = parser.parse_args()

    if not HAS_HYPER:
        print("Error: tableauhyperapi not installed.")
        print("  Run: pip install tableauhyperapi")
        sys.exit(1)

    # ── Resolve file list ──
    if args.file:
        filepath = Path(args.file)
        if not filepath.exists():
            print(f"Error: '{filepath}' not found."); sys.exit(1)
        if filepath.suffix.lower() != ".twbx":
            print("Error: only .twbx files are supported."); sys.exit(1)
        files   = [filepath]
        out_dir = filepath.parent
    else:
        folder = Path(args.folder)
        if not folder.exists():
            print(f"Error: '{folder}' not found."); sys.exit(1)
        files = sorted(folder.glob("*.twbx"))
        if not files:
            print(f"No .twbx files found in '{folder}'."); sys.exit(1)
        out_dir = folder
        print(f"\nFound {len(files)} workbook(s) in '{folder}'")

    # ── Process each workbook ──
    all_selected = []
    skipped      = []

    for fp in files:
        selected = process_workbook(fp)
        if selected:
            all_selected.extend(selected)
        else:
            skipped.append(fp.name)

    # ── Summary ──
    print(f"\n{'='*60}")
    print(f"  {len(all_selected)} figure(s) selected across {len(files) - len(skipped)} workbook(s)")
    if skipped:
        print(f"  Skipped ({len(skipped)}): {', '.join(skipped)}")
    print(f"{'='*60}")

    if not all_selected:
        print("\nNo figures selected — nothing to write.")
        sys.exit(0)

    print("\nWriting output...\n")

    label     = files[0].stem if len(files) == 1 else f"{len(files)} Workbooks"
    xlsx_path = out_dir / f"{args.output}.xlsx"
    html_path = out_dir / f"{args.output}.html"

    write_figures_xlsx(all_selected, label, xlsx_path)
    write_figures_html(all_selected, label, html_path)

    print(f"\nDone. Open {args.output}.html in any browser.")


if __name__ == "__main__":
    main()
